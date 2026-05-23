"""
市场推广Excel生成工具脚本
用法:
    python make_excel.py --type expense_detail --data data.json --output expense.xlsx
    python make_excel.py --type budget_tracker --data tracker.json --output budget.xlsx
"""

import argparse
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation


# ============================================================
# 集团样式
# ============================================================
HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
ALT_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
OVER_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GOOD_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

CENTER = Alignment(horizontal="center", vertical="center")
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

FMT_MONEY = '#,##0.00'
FMT_PCT = '0.0%'
FMT_INT = '#,##0'
FMT_DATE = 'YYYY-MM-DD'


def _write_header(ws, headers, row=1):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 30


def _write_rows(ws, rows, start_row=2):
    for ri, row in enumerate(rows):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=start_row + ri, column=ci, value=val)
            c.border = BORDER
            c.alignment = CENTER
            if ri % 2 == 1:
                c.fill = ALT_FILL


def _fmt_col(ws, col, start, end, fmt):
    for r in range(start, end + 1):
        ws[f"{col}{r}"].number_format = fmt


def _setup(ws):
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    for col in ws.columns:
        letter = col[0].column_letter
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[letter].width = min(max_len + 4, 45)


# ============================================================
# 各类型生成
# ============================================================
def generate_expense_detail(data, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = f"{data.get('month', '')}费用明细"

    headers = ["品牌", "区域", "费用类型", "预算金额", "实际执行", "执行率", "核销状态", "申请日期", "备注"]
    _write_header(ws, headers)
    _write_rows(ws, data.get("rows", []))

    lr = len(data.get("rows", [])) + 1
    _fmt_col(ws, "D", 2, lr, FMT_MONEY)
    _fmt_col(ws, "E", 2, lr, FMT_MONEY)
    _fmt_col(ws, "F", 2, lr, FMT_PCT)

    ws.conditional_formatting.add(
        f"F2:F{lr}",
        CellIsRule(operator="greaterThan", formula=["1"],
                   fill=OVER_FILL, font=Font(bold=True, color="9C0006"))
    )
    ws.conditional_formatting.add(
        f"F2:F{lr}",
        CellIsRule(operator="between", formula=["0.9", "1"],
                   fill=WARNING_FILL, font=Font(bold=True))
    )

    dv = DataValidation(type="list", formula1='"已核销,待核销,核销中,超期未核"', allow_blank=True)
    dv.add(f"G2:G{lr}")
    ws.add_data_validation(dv)

    _setup(ws)
    wb.save(output_path)
    return output_path


def generate_budget_tracker(data, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = f"{data.get('year', '')}预算追踪"

    headers = ["月份", "年度预算", "累计预算", "当月执行", "累计执行", "累计执行率", "剩余预算", "剩余占比", "预警"]
    _write_header(ws, headers)
    _write_rows(ws, data.get("rows", []))

    lr = len(data.get("rows", [])) + 1
    for cl in ["B", "C", "D", "E", "G"]:
        _fmt_col(ws, cl, 2, lr, FMT_MONEY)
    _fmt_col(ws, "F", 2, lr, FMT_PCT)
    _fmt_col(ws, "H", 2, lr, FMT_PCT)

    ws.conditional_formatting.add(
        f"F2:F{lr}",
        CellIsRule(operator="greaterThan", formula=["1"],
                   fill=OVER_FILL, font=Font(bold=True, color="9C0006"))
    )
    ws.conditional_formatting.add(
        f"F2:F{lr}",
        CellIsRule(operator="between", formula=["0.85", "1"],
                   fill=WARNING_FILL, font=Font(bold=True))
    )
    ws.conditional_formatting.add(
        f"H2:H{lr}",
        CellIsRule(operator="lessThan", formula=["0.15"],
                   fill=OVER_FILL, font=Font(bold=True, color="9C0006"))
    )

    _setup(ws)
    wb.save(output_path)
    return output_path


def generate_campaign_eval(data, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "活动效果评估"

    headers = ["活动名称", "品牌", "活动日期", "预算", "实际花费", "曝光量", "互动量", "线索数", "成交数", "CPL", "ROI", "达成率", "评分", "备注"]
    _write_header(ws, headers)
    _write_rows(ws, data.get("rows", []))

    lr = len(data.get("rows", [])) + 1
    for cl in ["D", "E"]:
        _fmt_col(ws, cl, 2, lr, FMT_MONEY)
    for cl in ["F", "G", "H", "I"]:
        _fmt_col(ws, cl, 2, lr, FMT_INT)
    _fmt_col(ws, "J", 2, lr, FMT_MONEY)
    _fmt_col(ws, "K", 2, lr, "0.00")
    _fmt_col(ws, "L", 2, lr, FMT_PCT)

    ws.conditional_formatting.add(
        f"K2:K{lr}",
        CellIsRule(operator="greaterThanOrEqual", formula=["3"],
                   fill=GOOD_FILL, font=Font(bold=True, color="006100"))
    )
    ws.conditional_formatting.add(
        f"K2:K{lr}",
        CellIsRule(operator="lessThan", formula=["1"],
                   fill=OVER_FILL, font=Font(bold=True, color="9C0006"))
    )
    ws.conditional_formatting.add(
        f"L2:L{lr}",
        CellIsRule(operator="greaterThanOrEqual", formula=["1"],
                   fill=GOOD_FILL, font=Font(bold=True, color="006100"))
    )

    _setup(ws)
    wb.save(output_path)
    return output_path


def generate_lead_tracker(data, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "线索转化追踪"

    headers = ["线索ID", "来源渠道", "品牌", "获取日期", "等级", "首跟日期", "跟进次数", "试驾日期", "成交日期", "转化周期(天)", "成交金额", "销售", "状态"]
    _write_header(ws, headers)
    _write_rows(ws, data.get("rows", []))

    lr = len(data.get("rows", [])) + 1
    _fmt_col(ws, "J", 2, lr, FMT_INT)
    _fmt_col(ws, "K", 2, lr, FMT_MONEY)

    dv = DataValidation(type="list", formula1='"待跟进,跟进中,已试驾,已成交,已战败"', allow_blank=True)
    dv.add(f"M2:M{lr}")
    ws.add_data_validation(dv)

    ws.conditional_formatting.add(
        f"J2:J{lr}",
        CellIsRule(operator="lessThanOrEqual", formula=["7"],
                   fill=GOOD_FILL, font=Font(bold=True, color="006100"))
    )

    _setup(ws)
    wb.save(output_path)
    return output_path


def generate_competitor_monitor(data, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "竞品动态监测"

    headers = ["日期", "竞品品牌", "活动类型", "活动主题", "活动区域", "预估费用(万)", "活动亮点", "效果评估", "对我方影响", "应对建议", "记录人"]
    _write_header(ws, headers)
    _write_rows(ws, data.get("rows", []))

    lr = len(data.get("rows", [])) + 1
    _fmt_col(ws, "F", 2, lr, FMT_MONEY)

    dv = DataValidation(type="list", formula1='"高,中,低"', allow_blank=True)
    dv.add(f"I2:I{lr}")
    ws.add_data_validation(dv)

    _setup(ws)
    wb.save(output_path)
    return output_path


# ============================================================
# 主流程
# ============================================================
GENERATORS = {
    "expense_detail": generate_expense_detail,
    "budget_tracker": generate_budget_tracker,
    "campaign_eval": generate_campaign_eval,
    "lead_tracker": generate_lead_tracker,
    "competitor_monitor": generate_competitor_monitor,
}


def main():
    parser = argparse.ArgumentParser(description="市场推广Excel生成工具")
    parser.add_argument("--type", dest="excel_type", default="expense_detail",
                        choices=list(GENERATORS.keys()),
                        help="表格类型")
    parser.add_argument("--data", help="JSON数据文件路径")
    parser.add_argument("--output", required=True, help="输出Excel文件路径")
    args = parser.parse_args()

    if args.data:
        with open(args.data, "r", encoding="utf-8") as f:
            data = json.load(f)
    else:
        data = {"rows": []}

    gen = GENERATORS[args.excel_type]
    output = gen(data, args.output)
    print(f"Excel 已生成: {output}")


if __name__ == "__main__":
    main()
