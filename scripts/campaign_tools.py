#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
活动执行工具脚本 - 为市场虾 Agent 提供市场活动全生命周期管理能力
支持四品牌活动的策划、执行、检查、改进全流程自动化
"""

import os
import sys
import json
import argparse
import csv
from pathlib import Path
from typing import List, Dict, Tuple, Optional, Union
from dataclasses import dataclass, field, asdict
from datetime import datetime, timedelta
from enum import Enum

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from pptx import Presentation
    from pptx.util import Inches, Pt, Emu
    from pptx.dml.color import RGBColor
    from pptx.enum.text import PP_ALIGN
    HAS_PPTX = True
except ImportError:
    HAS_PPTX = False


class EventType(Enum):
    NEW_CAR_LAUNCH = "new_car_launch"
    PROMOTION = "promotion"
    BRAND_EXPERIENCE = "brand_experience"
    CUSTOMER_APPRECIATION = "customer_appreciation"


class Brand(Enum):
    TRUMPCHI = "trumpchi"
    AUDI = "audi"
    HYPER = "hyper"
    AION = "aion"


BRAND_NAMES = {
    "trumpchi": "广汽传祺",
    "audi": "上汽奥迪",
    "hyper": "广汽昊铂",
    "aion": "广汽埃安",
}

EVENT_TYPE_NAMES = {
    "new_car_launch": "新车上市活动",
    "promotion": "促销活动",
    "brand_experience": "品牌体验活动",
    "customer_appreciation": "客户答谢活动",
}

# 四品牌活动差异化话术
BRAND_STRATEGY = {
    "trumpchi": {
        "theme": "驭见品质生活",
        "audience": "28-45岁家庭用户",
        "tone": "温馨、可靠、有温度",
        "channels": "社区社群、家庭类KOL、本地生活平台",
        "gift_direction": "实用家庭礼品（车载用品、家庭出游套装）",
        "key_points": "空间/油耗/安全/性价比",
    },
    "audi": {
        "theme": "尊享·非凡",
        "audience": "35-55岁高净值人群",
        "tone": "尊贵、私密、定制化",
        "channels": "高端生活方式媒体、私域社群、圈层传播",
        "gift_direction": "高端定制礼品（皮具、红酒、艺术周边）",
        "key_points": "品牌传承/豪华配置/专属服务",
    },
    "hyper": {
        "theme": "触碰未来",
        "audience": "25-40岁科技爱好者",
        "tone": "未来感、智能、先锋",
        "channels": "科技媒体、数码博主、新能源社群",
        "gift_direction": "科技类礼品（智能设备、品牌联名科技品）",
        "key_points": "智能驾驶/超充技术/OTA升级/科技配置",
    },
    "aion": {
        "theme": "年轻就要来电",
        "audience": "20-35岁年轻群体",
        "tone": "活力、年轻、潮流、环保",
        "channels": "小红书/抖音/B站、潮流KOL、校园社团",
        "gift_direction": "环保材质潮流礼品（再生材料周边、植物种子）",
        "key_points": "环保理念/超低能耗/潮流设计/智能互联",
    },
}


@dataclass
class CampaignConfig:
    """活动配置"""
    event_id: str = ""
    event_name: str = ""
    event_type: str = "new_car_launch"
    brand: str = "trumpchi"
    event_date: str = ""
    venue: str = ""
    venue_address: str = ""
    target_attendees: int = 100
    budget_total: float = 0.0
    budget_reserve: float = 0.0
    milestones: Dict[str, str] = field(default_factory=dict)


class CampaignTools:
    """活动执行工具主类"""

    DEFAULT_OUTPUT = Path("G:/openclaw/data/.openclaw/workspace/KaiQiJiTuan/output")
    DEFAULT_DATA = Path("G:/openclaw/data/.openclaw/workspace/KaiQiJiTuan/data")

    def __init__(self, output_dir: Optional[Path] = None):
        self.output_dir = Path(output_dir) if output_dir else self.DEFAULT_OUTPUT
        self.output_dir.mkdir(parents=True, exist_ok=True)

    # ========== 5.1 活动倒计时 ==========

    def generate_timeline(self, event_date: str, event_name: str,
                          event_type: str = "new_car_launch") -> List[Dict]:
        """生成活动时间线（倒推法）"""
        try:
            t0 = datetime.strptime(event_date, "%Y-%m-%d")
        except ValueError:
            raise ValueError("日期格式应为 YYYY-MM-DD")

        milestones_templates = {
            "new_car_launch": [
                (-45, "内部立项审批"),
                (-40, "确定活动方案、场地选址"),
                (-35, "供应商招标/比价"),
                (-30, "确定供应商、签订合同"),
                (-25, "设计物料初稿"),
                (-20, "物料定稿、进入制作"),
                (-15, "媒体/KOL邀约启动"),
                (-10, "物料到位检查、第一次彩排"),
                (-7, "客户/嘉宾邀约截止、第二次彩排"),
                (-3, "全场联排"),
                (-1, "物料进场、搭建完成"),
                (0, "【活动执行日】"),
                (1, "撤场"),
                (3, "效果数据汇总"),
                (7, "活动总结报告"),
            ],
            "promotion": [
                (-30, "确定促销政策"),
                (-25, "物料设计制作"),
                (-20, "广告投放上线"),
                (-15, "销售话术培训"),
                (-7, "展厅布置完成"),
                (-3, "全员动员会"),
                (0, "【促销活动开始】"),
                (30, "【促销活动结束】"),
                (33, "转化数据汇总"),
                (37, "活动效果复盘"),
            ],
            "brand_experience": [
                (-40, "体验区设计方案确定"),
                (-35, "技术供应商对接"),
                (-30, "互动环节开发完成"),
                (-25, "物料设计定稿"),
                (-15, "社交媒体预热"),
                (-7, "搭建进场"),
                (-3, "全流程联排"),
                (0, "【体验活动日】"),
                (3, "传播效果汇总"),
            ],
            "customer_appreciation": [
                (-30, "嘉宾名单筛选"),
                (-21, "邀请函发出"),
                (-14, "餐饮方案确认"),
                (-7, "礼品到位"),
                (-3, "场地布置检查"),
                (-1, "最终确认嘉宾出席"),
                (0, "【答谢活动日】"),
                (1, "感谢函发送"),
                (3, "客户回访"),
            ],
        }

        milestones = milestones_templates.get(event_type, milestones_templates["new_car_launch"])
        timeline = []
        for offset, task in milestones:
            target_date = t0 + timedelta(days=offset)
            delta = target_date - datetime.now()
            days_left = delta.days
            status = "已完成" if days_left < -1 else ("今天" if days_left == 0 else f"还有{days_left}天")
            timeline.append({
                "date": target_date.strftime("%Y-%m-%d"),
                "task": task,
                "offset": offset,
                "days_left": days_left,
                "status": status,
            })
        return timeline

    # ========== 5.2 检查清单 ==========

    def generate_checklist(self, event_type: str = "new_car_launch",
                           event_name: str = "") -> List[Dict]:
        """生成活动检查清单"""
        common_checks = [
            # 策划阶段
            {"phase": "策划阶段", "item": "活动方案审批通过", "category": "流程"},
            {"phase": "策划阶段", "item": "预算批复到位", "category": "财务"},
            {"phase": "策划阶段", "item": "场地合同签订", "category": "场地"},
            {"phase": "策划阶段", "item": "供应商合同签订", "category": "采购"},
            {"phase": "策划阶段", "item": "活动流程表定稿", "category": "策划"},
            # 执行阶段
            {"phase": "执行阶段", "item": "物料设计定稿", "category": "设计"},
            {"phase": "执行阶段", "item": "物料制作完成并验收入库", "category": "物料"},
            {"phase": "执行阶段", "item": "媒体/KOL邀请确认", "category": "媒体"},
            {"phase": "执行阶段", "item": "嘉宾邀请确认", "category": "接待"},
            {"phase": "执行阶段", "item": "产品展示车辆/展品就位", "category": "产品"},
            {"phase": "执行阶段", "item": "主持稿/新闻稿/发言稿定稿", "category": "文案"},
            {"phase": "执行阶段", "item": "礼品采购到位", "category": "物料"},
            {"phase": "执行阶段", "item": "全员分工表确认", "category": "人员"},
            {"phase": "执行阶段", "item": "第一次彩排完成", "category": "流程"},
            {"phase": "执行阶段", "item": "全场联排完成", "category": "流程"},
            # 现场阶段
            {"phase": "现场阶段", "item": "LED屏显示正常", "category": "设备"},
            {"phase": "现场阶段", "item": "音响系统调试完成", "category": "设备"},
            {"phase": "现场阶段", "item": "签到处布置完成", "category": "场地"},
            {"phase": "现场阶段", "item": "产品展示区布置完成", "category": "场地"},
            {"phase": "现场阶段", "item": "安全通道畅通", "category": "安全"},
            {"phase": "现场阶段", "item": "消防设施检查", "category": "安全"},
            {"phase": "现场阶段", "item": "应急医疗就位", "category": "安全"},
            # 收尾阶段
            {"phase": "收尾阶段", "item": "撤场完成", "category": "场地"},
            {"phase": "收尾阶段", "item": "费用结算完成", "category": "财务"},
            {"phase": "收尾阶段", "item": "照片/视频素材归档", "category": "归档"},
            {"phase": "收尾阶段", "item": "效果数据汇总", "category": "数据"},
            {"phase": "收尾阶段", "item": "活动总结报告完成", "category": "归档"},
        ]

        # 按活动类型添加专项检查
        type_specific = {
            "new_car_launch": [
                {"phase": "执行阶段", "item": "试驾路线规划完成", "category": "产品"},
                {"phase": "执行阶段", "item": "试驾保险购买", "category": "安全"},
                {"phase": "执行阶段", "item": "新闻通稿定稿", "category": "媒体"},
            ],
            "promotion": [
                {"phase": "执行阶段", "item": "价格政策文件确认", "category": "政策"},
                {"phase": "执行阶段", "item": "销售话术培训完成", "category": "培训"},
                {"phase": "执行阶段", "item": "CRM追踪表就绪", "category": "系统"},
            ],
            "brand_experience": [
                {"phase": "执行阶段", "item": "互动区设备调试完成", "category": "设备"},
                {"phase": "执行阶段", "item": "打卡引导物料到位", "category": "物料"},
                {"phase": "执行阶段", "item": "社交媒体预热发布", "category": "传播"},
            ],
            "customer_appreciation": [
                {"phase": "执行阶段", "item": "餐饮菜单确认", "category": "接待"},
                {"phase": "执行阶段", "item": "礼品包装完成", "category": "物料"},
                {"phase": "执行阶段", "item": "一对一接待分工确认", "category": "人员"},
            ],
        }

        checklist = common_checks.copy()
        for item in type_specific.get(event_type, []):
            checklist.append(item)

        return checklist

    # ========== 5.3 RACI 矩阵 ==========

    def generate_raci_matrix(self, event_type: str = "new_car_launch",
                            team_size: int = 8,
                            output_path: Optional[str] = None) -> Path:
        """生成人员分工 RACI 矩阵 Excel"""
        if not HAS_OPENPYXL:
            raise ImportError("需要安装 openpyxl: pip install openpyxl")

        roles = ["项目经理", "品牌经理", "设计师", "媒介经理", "执行导演",
                 "数据分析师", "客服经理", "销售经理",
                 "搭建方", "礼仪团队", "安保团队"][:team_size]

        tasks = {
            "new_car_launch": [
                ("整体统筹", "R", "A", "", "C", "I", "I", "I", "", "", ""),
                ("方案策划", "R", "A", "", "C", "C", "I", "", "", "", ""),
                ("物料设计", "C", "A", "R", "", "I", "", "", "", "", ""),
                ("媒体邀请", "I", "C", "", "R", "I", "", "", "", "", ""),
                ("场地搭建", "A", "I", "", "", "R", "", "", "R", "", ""),
                ("嘉宾接待", "I", "I", "", "", "I", "", "R", "", "", "R"),
                ("活动执行", "I", "", "", "", "R", "C", "", "", "R", ""),
                ("数据监控", "A", "", "", "", "", "R", "C", "I", "", ""),
                ("安全管控", "A", "", "", "", "C", "", "", "", "", "R"),
                ("费用结算", "R", "A", "", "", "", "C", "", "", "", ""),
            ],
            "promotion": [
                ("整体统筹", "R", "A", "", "C", "I", "I", "I", "", "", ""),
                ("价格政策", "C", "A", "", "", "", "C", "", "R", "", ""),
                ("物料准备", "C", "A", "R", "", "I", "", "", "", "", ""),
                ("广告投放", "I", "C", "", "R", "I", "C", "", "", "", ""),
                ("话术培训", "A", "I", "", "", "", "", "C", "R", "", ""),
                ("到店转化", "A", "", "", "", "", "", "R", "R", "", ""),
                ("数据追踪", "A", "", "", "", "", "R", "C", "C", "", ""),
            ],
            "brand_experience": [
                ("整体统筹", "R", "A", "", "C", "I", "I", "", "", "", ""),
                ("体验区设计", "C", "A", "R", "", "R", "", "", "", "", ""),
                ("互动开发", "C", "A", "R", "", "R", "", "", "", "", ""),
                ("传播预热", "I", "C", "", "R", "", "C", "", "", "", ""),
                ("现场执行", "I", "", "", "", "R", "C", "", "R", "R", ""),
                ("数据采集", "A", "", "", "", "", "R", "C", "", "", ""),
            ],
            "customer_appreciation": [
                ("整体统筹", "R", "A", "", "C", "I", "I", "", "", "", ""),
                ("嘉宾邀请", "C", "A", "", "", "", "", "R", "C", "", ""),
                ("礼品准备", "C", "A", "R", "", "", "", "", "", "", ""),
                ("餐饮安排", "A", "", "", "", "", "", "R", "", "", "R"),
                ("现场接待", "I", "I", "", "", "I", "", "R", "", "", "R"),
                ("关系维护", "A", "", "", "", "", "", "R", "R", "", ""),
            ],
        }

        task_list = tasks.get(event_type, tasks["new_car_launch"])

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RACI矩阵"

        # 样式
        header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        r_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        a_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))

        # 表头
        headers = ["任务"] + roles
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # 数据行
        for row_idx, (task_name, *r_vals) in enumerate(task_list, 2):
            ws.cell(row=row_idx, column=1, value=task_name).border = thin_border
            ws.cell(row=row_idx, column=1).font = Font(name="微软雅黑", size=10)
            for col_idx, val in enumerate(r_vals, 2):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                cell.font = Font(name="微软雅黑", size=10, bold=(val in ("R", "A")))
                if val == "R":
                    cell.fill = r_fill
                elif val == "A":
                    cell.fill = a_fill

        # 列宽
        ws.column_dimensions['A'].width = 16
        for col in range(2, len(roles) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 12

        # 保存
        if output_path is None:
            output_path = str(self.output_dir / f"raci_matrix_{event_type}.xlsx")
        wb.save(output_path)
        print(f"RACI矩阵已保存: {output_path}")
        return Path(output_path)

    # ========== 5.4 物料清单 ==========

    def generate_material_list(self, event_type: str = "new_car_launch",
                               brand: str = "trumpchi",
                               output_path: Optional[str] = None) -> Path:
        """生成物料清单 Excel"""
        if not HAS_OPENPYXL:
            raise ImportError("需要安装 openpyxl: pip install openpyxl")

        materials = {
            "new_car_launch": [
                ("主背景板", 1, "按场地定制", "喷绘+桁架", "", "", "T-1"),
                ("签到墙", 1, "2x3m", "喷绘+KT板", "", "", "T-1"),
                ("X展架", 8, "80x180cm", "PP纸+铝合金", "", "", "T-2"),
                ("易拉宝", 4, "80x200cm", "PVC海报", "", "", "T-2"),
                ("宣传册", 200, "A4 16P", "铜版纸157g", "", "", "T-3"),
                ("产品手册", 100, "A4 8P", "铜版纸200g", "", "", "T-3"),
                ("邀请函", 150, "A5对折", "特种纸+烫金", "", "", "T-14"),
                ("工作证", 50, "85x54mm", "PVC卡+挂绳", "", "", "T-3"),
                ("礼品袋", 200, "定制", "白卡纸+品牌印刷", "", "", "T-2"),
                ("礼品", 200, "定制", "", "", "", "T-2"),
                ("车贴", 6, "按车型", "PVC车贴", "", "", "T-1"),
                ("指引牌", 10, "60x40cm", "KT板", "", "", "T-1"),
                ("讲台装饰", 1, "", "品牌定制", "", "", "T-1"),
                ("矿泉水(品牌定制)", 200, "330ml", "", "", "", "T-1"),
            ],
            "promotion": [
                ("促销条幅", 4, "按场地", "灯布", "", "", "T-3"),
                ("价格标签", 20, "A5", "铜版纸", "", "", "T-3"),
                ("促销海报", 4, "80x60cm", "写真覆膜", "", "", "T-3"),
                ("礼品展示台", 2, "1.2x0.6m", "", "", "", "T-2"),
                ("礼品堆头", 1, "", "", "", "", "T-2"),
                ("气球装饰", 50, "", "", "", "", "T-1"),
                ("地贴指引", 5, "按需", "PVC地贴", "", "", "T-1"),
            ],
            "brand_experience": [
                ("体验区背景", 1, "按场地", "喷绘", "", "", "T-1"),
                ("互动设备", 3, "", "", "", "", "T-3"),
                ("打卡墙", 1, "3x2m", "定制+灯光", "", "", "T-1"),
                ("打卡道具", 20, "", "", "", "", "T-2"),
                ("引导地贴", 10, "", "PVC", "", "", "T-1"),
                ("社交媒体引导牌", 4, "30x20cm", "亚克力", "", "", "T-1"),
            ],
            "customer_appreciation": [
                ("欢迎背板", 1, "按场地", "喷绘", "", "", "T-1"),
                ("餐桌装饰", 15, "", "鲜花+台卡", "", "", "T-1"),
                ("席位卡", 100, "", "特种纸", "", "", "T-2"),
                ("伴手礼", 100, "定制", "品牌礼盒", "", "", "T-2"),
                ("感谢函", 100, "A5", "特种纸+烫金", "", "", "T-1"),
                ("签到簿", 1, "", "皮质+金箔", "", "", "T-1"),
            ],
        }

        items = materials.get(event_type, materials["new_car_launch"])
        brand_name = BRAND_NAMES.get(brand, brand)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "物料清单"

        header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')

        headers = ["序号", "物料名称", "数量", "规格", "材质", "制作方",
                   "单价", "总价", "到位时间", "状态"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        for idx, (name, qty, spec, material, maker, price, deadline) in enumerate(items, 1):
            row_data = [idx, name, qty, spec, material, maker, "", "", deadline, "待确认"]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=idx + 1, column=col, value=val)
                cell.font = Font(name="微软雅黑", size=10)
                cell.alignment = center_align
                cell.border = thin_border

        # 列宽
        col_widths = [6, 20, 8, 16, 16, 12, 10, 10, 12, 10]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        if output_path is None:
            output_path = str(self.output_dir /
                             f"material_list_{brand}_{event_type}.xlsx")
        wb.save(output_path)
        print(f"物料清单已保存: {output_path}")
        return Path(output_path)

    # ========== 5.5 风险矩阵 ==========

    def generate_risk_matrix(self, event_type: str = "new_car_launch",
                            output_path: Optional[str] = None) -> Path:
        """生成风险应对预案 Excel"""
        if not HAS_OPENPYXL:
            raise ImportError("需要安装 openpyxl: pip install openpyxl")

        risks = [
            ("天气风险", "户外活动遇雨", "中", "高",
             "准备雨棚、备用室内场地，提前3天关注天气预报",
             "场地负责人"),
            ("设备风险", "LED屏/音响故障", "低", "高",
             "备用设备一套、技术人员全程待命、彩排时充分测试",
             "技术负责人"),
            ("人员风险", "关键人员迟到/缺席", "中", "中",
             "准备备用主持人/发言人、提前1小时确认到达",
             "人事协调"),
            ("安全风险", "人员拥挤/消防隐患", "低", "极高",
             "安保团队现场管控、消防通道畅通、应急预案演练",
             "安全负责人"),
            ("舆情风险", "负面突发事件", "低", "极高",
             "公关团队待命、声明预案准备、媒体沟通机制",
             "公关负责人"),
            ("物料风险", "物料破损/丢失", "中", "中",
             "关键物料多备10%、现场修复工具包、备用打印",
             "物料负责人"),
            ("餐饮风险", "食品安全问题", "低", "极高",
             "选择有资质供应商、签订食品安全协议、全程监督",
             "接待负责人"),
            ("交通风险", "嘉宾/媒体迟到", "中", "低",
             "提前发送交通指南、安排接驳车、预留签到缓冲时间",
             "接待负责人"),
        ]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "风险应对预案"

        header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')
        wrap_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # 风险等级填充
        high_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        mid_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        low_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        headers = ["风险类型", "描述", "概率", "影响", "风险等级", "应对措施", "责任人"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        for idx, (r_type, desc, prob, impact, measure, person) in enumerate(risks, 2):
            # 计算风险等级
            prob_score = {"低": 1, "中": 2, "高": 3}.get(prob, 1)
            impact_score = {"低": 1, "中": 2, "高": 3, "极高": 4}.get(impact, 1)
            total = prob_score * impact_score
            if total <= 2:
                level = "低"
                level_fill = low_fill
            elif total <= 4:
                level = "中"
                level_fill = mid_fill
            else:
                level = "高"
                level_fill = high_fill

            row_data = [r_type, desc, prob, impact, level, measure, person]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=idx, column=col, value=val)
                cell.font = Font(name="微软雅黑", size=10)
                cell.border = thin_border
                if col in (1, 3, 4, 5, 7):
                    cell.alignment = center_align
                else:
                    cell.alignment = wrap_align
                if col == 5:
                    cell.fill = level_fill

        col_widths = [12, 20, 8, 8, 10, 50, 14]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.row_dimensions[1].height = 30
        for r in range(2, len(risks) + 2):
            ws.row_dimensions[r].height = 45

        if output_path is None:
            output_path = str(self.output_dir /
                             f"risk_matrix_{event_type}.xlsx")
        wb.save(output_path)
        print(f"风险矩阵已保存: {output_path}")
        return Path(output_path)

    # ========== 5.6 活动效果评估报告 ==========

    def generate_evaluation_report(self, event_id: str = "CAM-001",
                                  brand: str = "trumpchi",
                                  output_format: str = "xlsx",
                                  output_path: Optional[str] = None) -> Path:
        """生成活动效果评估报告"""
        brand_name = BRAND_NAMES.get(brand, brand)

        if output_format == "xlsx":
            return self._gen_eval_excel(event_id, brand_name, output_path)
        else:
            return self._gen_eval_markdown(event_id, brand_name, output_path)

    def _gen_eval_excel(self, event_id: str, brand_name: str,
                       output_path: Optional[str] = None) -> Path:
        if not HAS_OPENPYXL:
            raise ImportError("需要安装 openpyxl: pip install openpyxl")

        wb = openpyxl.Workbook()

        # Sheet 1: 活动概况
        ws1 = wb.active
        ws1.title = "活动概况"
        info = [
            ("活动编号", event_id),
            ("活动名称", f"{brand_name} - 市场活动"),
            ("活动日期", datetime.now().strftime("%Y-%m-%d")),
            ("品牌", brand_name),
        ]
        for row, (k, v) in enumerate(info, 1):
            ws1.cell(row=row, column=1, value=k).font = Font(bold=True)
            ws1.cell(row=row, column=2, value=v)

        # Sheet 2: KPI 达成
        ws2 = wb.create_sheet("KPI达成")
        kpi_headers = ["指标", "目标值", "实际值", "达成率", "评价"]
        for col, h in enumerate(kpi_headers, 1):
            ws2.cell(row=1, column=col, value=h).font = Font(bold=True)
        kpis = [
            ("到场人数", 100, 95, "95%", "接近目标"),
            ("媒体到场数", 20, 18, "90%", "达标"),
            ("直播在线人数", 50000, 45000, "90%", "达标"),
            ("有效线索数", 50, 42, "84%", "需优化"),
            ("线索有效率", "80%", "75%", "-5%", "需优化"),
            ("预算执行率", "100%", "92%", "92%", "良好"),
        ]
        for row, (k, t, a, rate, comment) in enumerate(kpis, 2):
            ws2.cell(row=row, column=1, value=k)
            ws2.cell(row=row, column=2, value=str(t))
            ws2.cell(row=row, column=3, value=str(a))
            ws2.cell(row=row, column=4, value=rate)
            ws2.cell(row=row, column=5, value=comment)

        # Sheet 3: 传播效果
        ws3 = wb.create_sheet("传播效果")
        media_headers = ["渠道", "曝光量", "互动量", "引流数", "CPM"]
        for col, h in enumerate(media_headers, 1):
            ws3.cell(row=1, column=col, value=h).font = Font(bold=True)

        # Sheet 4: 财务决算
        ws4 = wb.create_sheet("财务决算")
        fin_headers = ["类别", "预算", "实际", "差异", "差异率"]
        for col, h in enumerate(fin_headers, 1):
            ws4.cell(row=1, column=col, value=h).font = Font(bold=True)
        fin_items = [
            ("场地费", 20000, 18000, -2000, "-10%"),
            ("物料费", 30000, 32000, 2000, "+6.7%"),
            ("人员费", 15000, 14000, -1000, "-6.7%"),
            ("媒体费", 25000, 22000, -3000, "-12%"),
            ("餐饮费", 10000, 9500, -500, "-5%"),
            ("礼品费", 8000, 8500, 500, "+6.3%"),
            ("应急预备", 12000, 0, -12000, "-100%"),
            ("合计", 120000, 104000, -16000, "-13.3%"),
        ]
        for row, (cat, budget, actual, diff, rate) in enumerate(fin_items, 2):
            ws4.cell(row=row, column=1, value=cat)
            ws4.cell(row=row, column=2, value=budget)
            ws4.cell(row=row, column=3, value=actual)
            ws4.cell(row=row, column=4, value=diff)
            ws4.cell(row=row, column=5, value=rate)

        if output_path is None:
            output_path = str(self.output_dir / f"eval_report_{event_id}.xlsx")
        wb.save(output_path)
        print(f"效果评估报告已保存: {output_path}")
        return Path(output_path)

    def _gen_eval_markdown(self, event_id: str, brand_name: str,
                          output_path: Optional[str] = None) -> Path:
        now = datetime.now().strftime("%Y-%m-%d")
        md = f"""# 活动效果评估报告

## 活动概况
- 活动编号: {event_id}
- 活动名称: {brand_name} - 市场活动
- 报告日期: {now}
- 品牌: {brand_name}

## KPI 达成

| 指标 | 目标值 | 实际值 | 达成率 | 评价 |
|------|-------|-------|-------|------|
| 到场人数 | 100 | 95 | 95% | 接近目标 |
| 媒体到场数 | 20 | 18 | 90% | 达标 |
| 直播在线人数 | 50,000 | 45,000 | 90% | 达标 |
| 有效线索数 | 50 | 42 | 84% | 需优化 |

## 财务决算

| 类别 | 预算 | 实际 | 差异 |
|------|------|------|------|
| 场地费 | 20,000 | 18,000 | -2,000 |
| 物料费 | 30,000 | 32,000 | +2,000 |
| 媒体费 | 25,000 | 22,000 | -3,000 |
| **合计** | **120,000** | **104,000** | **-16,000** |

## 亮点复盘
1. 待填写
2. 待填写

## 问题复盘
1. 待填写
2. 待填写

## 改进建议
1. 待填写
2. 待填写
"""
        if output_path is None:
            output_path = str(self.output_dir / f"eval_report_{event_id}.md")
        Path(output_path).write_text(md, encoding='utf-8')
        print(f"效果评估报告已保存: {output_path}")
        return Path(output_path)

    # ========== 品牌策略查询 ==========

    def get_brand_strategy(self, brand: str) -> Dict:
        """获取品牌活动差异化策略"""
        if brand not in BRAND_STRATEGY:
            raise ValueError(f"未知品牌: {brand}, 可选: {list(BRAND_STRATEGY.keys())}")
        return BRAND_STRATEGY[brand]


def main():
    parser = argparse.ArgumentParser(description="活动执行工具 - 市场虾 Agent")
    sub = parser.add_subparsers(dest="command")

    # timeline
    sp = sub.add_parser("timeline", help="生成活动时间线")
    sp.add_argument("--date", required=True, help="活动日期 YYYY-MM-DD")
    sp.add_argument("--name", default="市场活动", help="活动名称")
    sp.add_argument("--type", default="new_car_launch",
                    choices=list(EVENT_TYPE_NAMES.keys()), help="活动类型")

    # checklist
    sp = sub.add_parser("checklist", help="生成检查清单")
    sp.add_argument("--type", default="new_car_launch",
                    choices=list(EVENT_TYPE_NAMES.keys()))
    sp.add_argument("--name", default="", help="活动名称")

    # raci
    sp = sub.add_parser("raci", help="生成RACI矩阵")
    sp.add_argument("--type", default="new_car_launch",
                    choices=list(EVENT_TYPE_NAMES.keys()))
    sp.add_argument("--team-size", type=int, default=8)
    sp.add_argument("--output", default=None)

    # materials
    sp = sub.add_parser("materials", help="生成物料清单")
    sp.add_argument("--type", default="new_car_launch",
                    choices=list(EVENT_TYPE_NAMES.keys()))
    sp.add_argument("--brand", default="trumpchi",
                    choices=["trumpchi", "audi", "hyper", "aion"])
    sp.add_argument("--output", default=None)

    # risk
    sp = sub.add_parser("risk", help="生成风险应对预案")
    sp.add_argument("--type", default="new_car_launch",
                    choices=list(EVENT_TYPE_NAMES.keys()))
    sp.add_argument("--output", default=None)

    # analyze
    sp = sub.add_parser("analyze", help="生成活动效果评估报告")
    sp.add_argument("--event-id", default="CAM-001")
    sp.add_argument("--brand", default="trumpchi",
                    choices=["trumpchi", "audi", "hyper", "aion"])
    sp.add_argument("--format", default="xlsx", choices=["xlsx", "md"])
    sp.add_argument("--output", default=None)

    # strategy
    sp = sub.add_parser("strategy", help="查看品牌活动策略")
    sp.add_argument("--brand", required=True,
                    choices=["trumpchi", "audi", "hyper", "aion"])

    args = parser.parse_args()
    tools = CampaignTools()

    if args.command == "timeline":
        timeline = tools.generate_timeline(args.date, args.name, args.type)
        print(f"\n=== {args.name} 时间线 ===")
        for t in timeline:
            print(f"  {t['date']}  [{t['status']}]  {t['task']}")

    elif args.command == "checklist":
        checklist = tools.generate_checklist(args.type, args.name)
        print(f"\n=== {EVENT_TYPE_NAMES.get(args.type, args.type)} 检查清单 ===\n")
        current_phase = ""
        for item in checklist:
            if item["phase"] != current_phase:
                current_phase = item["phase"]
                print(f"\n--- {current_phase} ---")
            print(f"  [{item['category']}] {item['item']}")

    elif args.command == "raci":
        tools.generate_raci_matrix(args.type, args.team_size, args.output)

    elif args.command == "materials":
        tools.generate_material_list(args.type, args.brand, args.output)

    elif args.command == "risk":
        tools.generate_risk_matrix(args.type, args.output)

    elif args.command == "analyze":
        tools.generate_evaluation_report(args.event_id, args.brand,
                                        args.format, args.output)

    elif args.command == "strategy":
        strategy = tools.get_brand_strategy(args.brand)
        print(f"\n=== {BRAND_NAMES[args.brand]} 活动策略 ===")
        for k, v in strategy.items():
            print(f"  {k}: {v}")


if __name__ == "__main__":
    main()